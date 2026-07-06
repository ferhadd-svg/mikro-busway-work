import re
from datetime import date

import openpyxl

from app.schemas.boq import BOQLineItem, BOQRun, FlagAnswers
from app.services.quotation_builder import (
    _fill_template,
    _insert_row,
    _refresh_lme_mentions,
)


class FakeSalesperson:
    name = "Eric Tan"
    title = "Sales Manager"
    mobile = "+60 12-000 0000"
    email = "eric@mikro.com.my"


def _runs():
    return [BOQRun(
        run_id="RUN-1",
        routing="FROM TX TO MSB",
        run_type="TX-MSB",
        material="AL",
        frame_rating_a=630,
        earth_pct=50,
        items=[
            BOQLineItem(description="FEEDER C/W INTEGRAL EARTH 500A (630A) 3P4W+50%E (ALUMINIUM)",
                        unit="m", qty=10, unit_rate_myr=100.0, amount_myr=1000.0),
            BOQLineItem(description="END CLOSURE (630A)", unit="No.", qty=1,
                        unit_rate_myr=500.0, amount_myr=500.0),
        ],
    )]


def _flags():
    return FlagAnswers(lme_usd_per_mt=2600.0, usd_to_myr=4.4500)


def _find_row(ws, predicate):
    for row in ws.iter_rows():
        for c in row:
            if predicate(c.value):
                return c.row
    return None


def _template():
    """Minimal <<token>>-style template: item-table header at row 3
    (amounts in F), merged totals labels, merged remarks footer."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Our Ref: <<OUR_REF>>"
    for col, header in enumerate(
            ["No.", "DESCRIPTION", "UNIT", "QTY", "UNIT RATE (RM)", "AMOUNT (RM)"], 1):
        ws.cell(row=3, column=col, value=header)
    ws.merge_cells("A5:D5"); ws["A5"] = "SUB-TOTAL (RM)"
    ws.merge_cells("A6:D6"); ws["A6"] = "10% SST (RM)"
    ws.merge_cells("A7:D7"); ws["A7"] = "GRAND TOTAL (RM)"
    ws.merge_cells("A9:F9")
    ws["A9"] = "1. The quantities quoted are rough estimation only. Final busway amount shall based on the actual delivery."
    return wb, ws


def test_insert_row_keeps_merges_aligned_with_values():
    wb, ws = _template()
    _insert_row(ws, 4)
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert ws["A6"].value == "SUB-TOTAL (RM)"
    assert "A6:D6" in merged
    assert "A5:D5" not in merged


def test_fill_template_totals_and_merges():
    wb, ws = _template()
    _fill_template(ws, _runs(), _flags(), FakeSalesperson(),
                   our_ref="MK/Q/001", client_name="ACME Sdn Bhd",
                   attn=None, me_consultant=None)

    # Token replaced
    assert ws["A1"].value == "Our Ref: MK/Q/001"

    # Labels still sit at the anchor of their (shifted) merged ranges
    merged = {str(r) for r in ws.merged_cells.ranges}
    label_rows = {}
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or "")
            if v.startswith(("SUB-TOTAL", "10% SST", "GRAND TOTAL")):
                label_rows[v.split(" (")[0]] = cell.row
                assert f"A{cell.row}:D{cell.row}" in merged, \
                    f"merge for '{v}' not aligned with its row {cell.row}"

    # Totals wired as formulas in the AMOUNT column (F): sub-total sums the
    # per-run sum row directly above it, SST and grand chain off it.
    sub_r, sst_r, grand_r = (label_rows["SUB-TOTAL"], label_rows["10% SST"],
                             label_rows["GRAND TOTAL"])
    assert ws.cell(row=sub_r, column=6).value == f"=SUM(F{sub_r - 1})"
    assert ws.cell(row=sst_r, column=6).value == f"=F{sub_r}*10%"
    assert ws.cell(row=grand_r, column=6).value == f"=SUM(F{sub_r},F{sst_r})"

    # Component line: shortened description, =QTY*RATE formula in F
    feeder_row = _find_row(ws, lambda v: v == "FEEDER C/W INTEGRAL EARTH")
    assert feeder_row is not None
    assert ws.cell(row=feeder_row, column=6).value == f"=D{feeder_row}*E{feeder_row}"


def _eric_style_template():
    """Mirror of the real Mikro salesperson template: labelled header fields
    (no <<tokens>>), Quantity/Unit/Unit Rate/Total Amount in E/F/G/H, sample
    items with formulas, totals labels in F/G/B."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["D1"] = "MIKRO BUSWAY SDN BHD (201001002004)"
    ws["A8"] = "To"; ws["B8"] = ": "
    ws["A9"] = "Attn"; ws["B9"] = ": "
    ws["E8"] = "Date"; ws["F8"] = ": "
    ws.merge_cells("A12:H12"); ws["A12"] = "QUOTATION"
    ws["A14"] = "OUR REF    : "
    ws["G14"] = "M&E : "
    ws["A16"] = "No."; ws.merge_cells("B16:D16"); ws["B16"] = "Description"
    ws["E16"] = "Quantity"; ws["F16"] = "Unit"
    ws["G16"] = "Unit Rate"; ws["H16"] = "Total Amount"
    ws["H17"] = "Ex-Nilai Factory in RM"
    # Sample items shipped with the template
    ws["A18"] = 1; ws["B18"] = "MIKRO BUSWAY # 2500A SAMPLE"
    ws["B19"] = "FEEDER C/W INTEGRAL EARTH"; ws["F19"] = "MTR"; ws["H19"] = "=E19*G19"
    ws["H20"] = "=SUM(H19:H19)"
    ws.merge_cells("F22:G22"); ws["F22"] = "Sub-Total Amount  :"
    ws["H22"] = "=SUM(H20)"
    ws["G23"] = "10% SST  :"; ws["H23"] = "=SUM(H22*10%)"
    ws.merge_cells("B24:G24")
    ws["B24"] = "Grand Total Amount for Item No. 1 to 2 , Ex-Nilai Factory in RM  :"
    ws["H24"] = "=SUM(H22,H23)"
    ws["B26"] = "Remarks :"
    ws["B27"] = "1. The quantities quoted are rough estimation only. Final busway amount shall based on the actual"
    return wb, ws


def test_fill_real_style_template():
    wb, ws = _eric_style_template()
    _fill_template(ws, _runs(), _flags(), FakeSalesperson(),
                   our_ref="MK/Q/2026/001", client_name="ACME Sdn Bhd",
                   attn="Mr. Lim", me_consultant="XYZ Consult")

    # Labelled header fields filled (this template has no <<tokens>>)
    assert ws["B8"].value == ": ACME Sdn Bhd"
    assert ws["B9"].value == ": Mr. Lim"
    assert ws["A14"].value.endswith("MK/Q/2026/001")
    assert ws["G14"].value.endswith("XYZ Consult")
    assert re.fullmatch(r": \d{2}-\d{2}-\d{4}", str(ws["F8"].value))  # ": <date>" filled

    # Sample items removed
    all_text = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert not any("SAMPLE" in t for t in all_text)

    # House format: run number in A + spec title in B, routing on next row
    title_row = _find_row(ws, lambda v: isinstance(v, str) and v.startswith("MIKRO BUSWAY #"))
    assert title_row is not None
    assert ws.cell(row=title_row, column=1).value == 1
    assert ws.cell(row=title_row, column=2).value == \
        "MIKRO BUSWAY # 630A TPNE, 3P4W+50%E, 600VAC, 50Hz (ALUMINIUM) - IP54"
    assert ws.cell(row=title_row + 1, column=2).value == "FROM TX TO MSB"

    # Component line in the template's own columns with =E*G formula,
    # description shortened (title already carries the spec)
    feeder_row = _find_row(ws, lambda v: v == "FEEDER C/W INTEGRAL EARTH")
    assert ws.cell(row=feeder_row, column=5).value == 10       # Quantity (E)
    assert ws.cell(row=feeder_row, column=6).value == "m"      # Unit (F)
    assert ws.cell(row=feeder_row, column=7).value == 100.0    # Unit Rate (G)
    assert ws.cell(row=feeder_row, column=8).value == f"=E{feeder_row}*G{feeder_row}"

    # Per-run sum row directly under the run's last component
    end_closure_row = _find_row(ws, lambda v: v == "END CLOSURE")
    sum_row = end_closure_row + 1
    assert ws.cell(row=sum_row, column=8).value == f"=SUM(H{feeder_row}:H{end_closure_row})"

    # Totals chained off the per-run sums; merges still aligned
    merged = {str(r) for r in ws.merged_cells.ranges}
    sub_row = _find_row(ws, lambda v: str(v or "").startswith("Sub-Total"))
    sst_row = _find_row(ws, lambda v: str(v or "").startswith("10% SST"))
    grand_row = _find_row(ws, lambda v: str(v or "").startswith("Grand Total"))
    assert ws.cell(row=sub_row, column=8).value == f"=SUM(H{sum_row})"
    assert ws.cell(row=sst_row, column=8).value == f"=H{sub_row}*10%"
    assert ws.cell(row=grand_row, column=8).value == f"=SUM(H{sub_row},H{sst_row})"
    assert f"F{sub_row}:G{sub_row}" in merged
    assert f"B{grand_row}:G{grand_row}" in merged

    # Grand-total label refreshed for the actual run count (1 run here)
    grand_label = ws.cell(row=grand_row, column=2).value
    assert "Item No. 1 ," in grand_label
    assert "1 to 2" not in grand_label

    # Remarks survive below the totals
    assert any(str(c.value or "").startswith("Remarks")
               for row in ws.iter_rows() for c in row)


def _add_lme_remarks(ws):
    """Static remark/validity text as shipped in the real Eric template:
    aluminium block in B/C, copper block in L/M, with the previous
    quotation's LME rate and date hard-typed."""
    ws["B28"] = "Validity"
    ws["C28"] = ": Based on LME Alu.@USD3,666/MT and thereafter depands on current LME Aluminium price."
    ws["B29"] = "3. The prices quoted are based on 05-06-2026 LME Aluminium @USD3,666/MT."
    ws["M28"] = ": 3 days from date of quotation. Thereafter shall depands on LME Copper price."
    ws["M29"] = ": 20% Deposit is required to secure LME Copper @ USD 13, 690/MT. "
    ws["L30"] = "3.  The prices quoted are based on 18-06-2026 LME Copper @ USD 13,690/MT."
    ws["L31"] = "NOTE : The prices quoted are based on Exchange Rate of USD 1.00 = RM 4.13"


def test_refresh_lme_mentions_rewrites_rate_and_date():
    wb = openpyxl.Workbook()
    ws = wb.active
    _add_lme_remarks(ws)
    _refresh_lme_mentions(ws, _flags())
    today = date.today().strftime("%d-%m-%Y")

    # Aluminium block: rate swapped in-place, template's own spacing kept
    assert ws["C28"].value == \
        ": Based on LME Alu.@USD2,600/MT and thereafter depands on current LME Aluminium price."
    assert ws["B29"].value == \
        f"3. The prices quoted are based on {today} LME Aluminium @USD2,600/MT."

    # Copper block: spaced style kept, stray space inside "13, 690" cleaned
    assert ws["M29"].value == ": 20% Deposit is required to secure LME Copper @ USD 2,600/MT. "
    assert ws["L30"].value == \
        f"3.  The prices quoted are based on {today} LME Copper @ USD 2,600/MT."

    # LME mention without a rate or date stays untouched
    assert ws["M28"].value == \
        ": 3 days from date of quotation. Thereafter shall depands on LME Copper price."
    # Non-LME cell with a USD figure stays untouched
    assert ws["L31"].value == \
        "NOTE : The prices quoted are based on Exchange Rate of USD 1.00 = RM 4.13"


def test_fill_template_refreshes_lme_remarks():
    wb, ws = _eric_style_template()
    _add_lme_remarks(ws)
    _fill_template(ws, _runs(), _flags(), FakeSalesperson(),
                   our_ref="X", client_name="Y", attn=None, me_consultant=None)
    today = date.today().strftime("%d-%m-%Y")

    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert f"3. The prices quoted are based on {today} LME Aluminium @USD2,600/MT." in texts
    assert f"3.  The prices quoted are based on {today} LME Copper @ USD 2,600/MT." in texts
    assert ": Based on LME Alu.@USD2,600/MT and thereafter depands on current LME Aluminium price." in texts
    # No stale rate or date survives anywhere in the sheet
    assert not any("3,666" in t or "690" in t or "05-06-2026" in t or "18-06-2026" in t
                   for t in texts)
    # Non-LME exchange-rate note untouched
    assert "NOTE : The prices quoted are based on Exchange Rate of USD 1.00 = RM 4.13" in texts


def test_multiple_runs_all_land_above_subtotal():
    wb, ws = _eric_style_template()
    runs = [_runs()[0],
            BOQRun(run_id="RUN-2", routing="FROM MSB TO L5", run_type="MSB-Riser",
                   material="CU", frame_rating_a=800, earth_pct=100,
                   items=[BOQLineItem(description="FEEDER C/W INTEGRAL EARTH 800A",
                                      unit="m", qty=20, unit_rate_myr=200.0,
                                      amount_myr=4000.0)],
                   piu_items=[BOQLineItem(description="250A TPN MCCB (HYUNDAI)",
                                          unit="No.", qty=2, unit_rate_myr=3000.0,
                                          amount_myr=6000.0)])]
    _fill_template(ws, runs, _flags(), FakeSalesperson(),
                   our_ref="X", client_name="Y", attn=None, me_consultant=None)

    sub_row = _find_row(ws, lambda v: str(v or "").startswith("Sub-Total"))
    riser_title_row = _find_row(
        ws, lambda v: isinstance(v, str) and v.startswith("MIKRO BUSWAY # 800A"))
    assert riser_title_row is not None
    assert riser_title_row < sub_row, "second run leaked below the SUB-TOTAL row"
    assert ws.cell(row=riser_title_row, column=1).value == 2  # numbered as item 2

    # Sub-total sums both per-run sum rows
    sub_formula = str(ws.cell(row=sub_row, column=8).value)
    assert sub_formula.startswith("=SUM(") and sub_formula.count("H") == 2

    # Grand-total label refreshed to "1 to 2"
    grand_row = _find_row(ws, lambda v: str(v or "").startswith("Grand Total"))
    assert "Item No. 1 to 2" in ws.cell(row=grand_row, column=2).value

    # PIU section label present within the run block
    piu_row = _find_row(ws, lambda v: v == "PLUG-IN UNITS (PIU) :")
    assert piu_row is not None and piu_row < sub_row
