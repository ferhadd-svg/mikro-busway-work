"""
Parse the Mikro Busway price list Excel (.xls/.xlsx).

Expected sheets: "List Price (Al)", "List Price (Cu)", "PIU", "BI METAL PLATE"

All rates used as-is (no multipliers); rounded to nearest RM at quotation time.
"""

import xlrd
import openpyxl
from pathlib import Path
from typing import Optional
import json


# Frame rating ladder — maps nominal to Mikro frame rating
FRAME_LADDER = [200, 400, 630, 800, 1000, 1250, 1600, 2000, 2500, 3200, 4000, 5000]

# Reverse of the parser's LABEL_MAP key prefixes (excluding "feeder", handled
# specially in all_rates() since it has an extra earth-% suffix segment) —
# maps a stored key prefix back to a human-readable category for display.
_CATEGORY_PREFIX_MAP = {
    "flange_end_box": "Flange End Box",
    "flange_end": "Flange End",
    "elbow": "Elbow",
    "flexible": "Flexible Conductor",
    "hanger_clamp": "Mounting Clamp",
    "end_closure": "End Closure",
    "fixed_hanger": "Fixed Hanger",
    "spring_hanger": "Spring Hanger",
    "plugin_hole": "Plug-in Opening",
}


def _strip_frame_suffix(key: str) -> str:
    """'elbow_800' -> 'elbow'; 'flange_end_box_630' -> 'flange_end_box'."""
    parts = key.rsplit("_", 1)
    return parts[0] if len(parts) == 2 else key


def _extract_trailing_int(key: str) -> Optional[int]:
    parts = key.rsplit("_", 1)
    if len(parts) == 2:
        try:
            return int(parts[1])
        except ValueError:
            return None
    return None


def resolve_frame_rating(nominal_a: int) -> int:
    """Return the Mikro frame rating for a given nominal amperage.
    E.g. 500 → 630, 100 → 200 (minimum frame = 200A).
    """
    for frame in FRAME_LADDER:
        if nominal_a <= frame:
            return frame
    return FRAME_LADDER[-1]


class PriceList:
    """In-memory price lookup loaded from the current price list file."""

    def __init__(self):
        self._al: dict = {}
        self._cu: dict = {}
        self._piu: dict = {}
        self._bimetal: dict = {}
        self._bimetal_dims: dict = {}   # frame_a -> (no, w_mm, l_mm)
        self._loaded_file: Optional[str] = None

    def load(self, path: Path) -> None:
        suffix = path.suffix.lower()
        if suffix == ".xls":
            self._load_xls(path)
        else:
            self._load_xlsx(path)
        self._loaded_file = str(path)

    # ------------------------------------------------------------------ #
    #  Public lookup API                                                   #
    # ------------------------------------------------------------------ #

    def feeder(self, frame_a: int, earth_pct: int, material: str) -> float:
        """Feeder 3P(4W) + earth% rate."""
        store = self._cu if material == "CU" else self._al
        key = f"feeder_{frame_a}_{earth_pct}"
        return store.get(key, 0.0)

    def flange_end(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"flange_end_{frame_a}", 0.0)

    def flange_end_box(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"flange_end_box_{frame_a}", 0.0)

    def cable_entry_box(self, frame_a: int, material: str) -> float:
        """Sum of Flange End + Flange End Box."""
        return self.flange_end(frame_a, material) + self.flange_end_box(frame_a, material)

    def elbow(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"elbow_{frame_a}", 0.0)

    def vertical_elbow(self, frame_a: int, material: str) -> float:
        return self.elbow(frame_a, material) + 200.0

    def flexible_conductor(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"flexible_{frame_a}", 0.0)

    def mounting_clamp(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"hanger_clamp_{frame_a}", 0.0)

    def end_closure(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"end_closure_{frame_a}", 0.0)

    def fixed_hanger(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"fixed_hanger_{frame_a}", store.get("fixed_hanger_default", 40.0))

    def spring_hanger(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"spring_hanger_{frame_a}", store.get("spring_hanger_default", 60.0))

    def plugin_opening(self, frame_a: int, material: str) -> float:
        store = self._cu if material == "CU" else self._al
        return store.get(f"plugin_hole_{frame_a}", 0.0)

    def bimetal(self, frame_a: int) -> float:
        """Bi-metal plate — aluminium runs only."""
        return self._bimetal.get(f"bimetal_{frame_a}", 0.0)

    def bimetal_dims(self, frame_a: int):
        """(No, W_mm, L_mm) for the bi-metal plate at this frame, or None."""
        return self._bimetal_dims.get(frame_a)

    def piu(self, rating_a: int, ka: int = 26) -> float:
        """PIU Hyundai MCCB rate by ampere + kA tier."""
        key = f"piu_{rating_a}_{ka}"
        if key in self._piu:
            return self._piu[key]
        # Fallback: match nearest band
        for band_a in sorted(self._piu_bands(ka)):
            if rating_a <= band_a:
                return self._piu.get(f"piu_{band_a}_{ka}", 0.0)
        return 0.0

    def piu_bands(self, ka: int = 26) -> list[int]:
        return self._piu_bands(ka)

    def _piu_bands(self, ka: int) -> list[int]:
        prefix = f"piu_"
        suffix = f"_{ka}"
        bands = []
        for k in self._piu:
            if k.startswith(prefix) and k.endswith(suffix):
                try:
                    bands.append(int(k[len(prefix):-len(suffix)]))
                except ValueError:
                    pass
        return bands

    def is_loaded(self) -> bool:
        return self._loaded_file is not None

    def loaded_file(self) -> Optional[str]:
        return self._loaded_file

    def all_rates(self) -> list[dict]:
        """Flatten _al/_cu/_piu/_bimetal into structured rows for display
        only — NOT used by any lookup method above, so boq_builder.py and
        the /lookup endpoints are unaffected by this method's behavior."""
        rows: list[dict] = []
        for material, store in (("AL", self._al), ("CU", self._cu)):
            for key, rate in store.items():
                if key.startswith("feeder_"):
                    parts = key.split("_")
                    if len(parts) == 3:
                        rows.append({
                            "category": "Feeder", "material": material,
                            "frame_a": int(parts[1]), "earth_pct": int(parts[2]),
                            "rate": rate,
                        })
                    continue
                cat = _CATEGORY_PREFIX_MAP.get(_strip_frame_suffix(key))
                if cat:
                    rows.append({
                        "category": cat, "material": material,
                        "frame_a": _extract_trailing_int(key), "rate": rate,
                    })

        for key, rate in self._piu.items():
            parts = key.split("_")
            if len(parts) == 3:
                rows.append({
                    "category": "PIU", "rating_a": int(parts[1]), "ka": int(parts[2]),
                    "rate": rate,
                })

        for key, rate in self._bimetal.items():
            parts = key.split("_")
            if len(parts) == 2:
                rows.append({"category": "Bi-Metal Plate", "frame_a": int(parts[1]), "rate": rate})

        return rows

    # ------------------------------------------------------------------ #
    #  Internal parsers — xls / xlsx                                      #
    # ------------------------------------------------------------------ #

    def _load_xls(self, path: Path) -> None:
        wb = xlrd.open_workbook(str(path))
        for sheet_name in wb.sheet_names():
            sheet = wb.sheet_by_name(sheet_name)
            sn = sheet_name.strip().lower()
            if "al" in sn and "list" in sn:
                self._al = _parse_price_sheet_xls(sheet)
            elif "cu" in sn and "list" in sn:
                self._cu = _parse_price_sheet_xls(sheet)
            elif "piu" in sn:
                self._piu = _parse_piu_sheet_xls(sheet)
            elif "bi metal" in sn or "bi-metal" in sn or "bimetal" in sn:
                self._bimetal, self._bimetal_dims = _parse_bimetal_sheet_xls(sheet)

    def _load_xlsx(self, path: Path) -> None:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sn = sheet_name.strip().lower()
            if "al" in sn and "list" in sn:
                self._al = _parse_price_sheet_xlsx(ws)
            elif "cu" in sn and "list" in sn:
                self._cu = _parse_price_sheet_xlsx(ws)
            elif "piu" in sn:
                self._piu = _parse_piu_sheet_xlsx(ws)
            elif "bi metal" in sn or "bi-metal" in sn or "bimetal" in sn:
                self._bimetal, self._bimetal_dims = _parse_bimetal_sheet_xlsx(ws)


# ------------------------------------------------------------------ #
#  Sheet parsers — build {key: rate} dicts                            #
# ------------------------------------------------------------------ #
#
#  These parsers scan row-by-row looking for known row labels and
#  column headers that contain frame ratings.  Cells containing
#  numeric values are stored under deterministic keys like
#  "feeder_630_50" or "elbow_800".
#
#  Because price-list layouts can shift, each parser is written
#  defensively: it logs warnings but never raises.

def _rows_xls(sheet):
    for r in range(sheet.nrows):
        yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]


def _rows_xlsx(ws):
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def _parse_price_sheet_xls(sheet) -> dict:
    return _parse_price_sheet(list(_rows_xls(sheet)))


def _parse_price_sheet_xlsx(ws) -> dict:
    return _parse_price_sheet(list(_rows_xlsx(ws)))


def _parse_piu_sheet_xls(sheet) -> dict:
    return _parse_piu_sheet(list(_rows_xls(sheet)))


def _parse_piu_sheet_xlsx(ws) -> dict:
    return _parse_piu_sheet(list(_rows_xlsx(ws)))


def _parse_bimetal_sheet_xls(sheet) -> dict:
    return _parse_bimetal_sheet(list(_rows_xls(sheet)))


def _parse_bimetal_sheet_xlsx(ws) -> dict:
    return _parse_bimetal_sheet(list(_rows_xlsx(ws)))


def _to_float(v) -> Optional[float]:
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _extract_amperage(value) -> Optional[int]:
    """Extract a single amperage value from a header/label cell. Handles
    text cells ("400A") as well as numeric cells that Excel/xlrd/openpyxl
    hand back as plain int/float (e.g. 800.0) with no "A" suffix at all —
    str(800.0) == "800.0" fails both text regexes below because of the
    decimal point, so numeric cells must be checked directly first."""
    import re
    if isinstance(value, (int, float)):
        v = int(round(value))
        return v if 100 <= v <= 9999 else None
    text = str(value)
    m = re.search(r"(\d{2,5})\s*[Aa]", text)
    if m:
        return int(m.group(1))
    m = re.search(r"^\s*(\d{2,5})\s*$", text)
    if m:
        v = int(m.group(1))
        if 100 <= v <= 9999:
            return v
    return None


def _extract_upper_amperage(value) -> Optional[int]:
    """For range-style row labels like "32A - 100A" returns the upper bound
    (100) — the correct key for PriceList.piu()'s "nearest band, rounded up"
    fallback lookup. For a single label like "630A (c/w busbar)" returns
    that value. Falls back to _extract_amperage's numeric-cell handling."""
    import re
    if isinstance(value, (int, float)):
        return _extract_amperage(value)
    matches = re.findall(r"(\d{2,5})\s*[Aa]", str(value))
    return int(matches[-1]) if matches else None


def _parse_price_sheet(rows: list) -> dict:
    """
    Generic parser for List Price (Al) / List Price (Cu).
    Strategy: scan headers row for ampere columns, then match row labels.
    """
    result: dict = {}
    # Find header row: contains multiple amperage values
    header_row_idx = None
    frame_cols: dict[int, int] = {}  # col_index -> frame_a

    for i, row in enumerate(rows):
        found = []
        for j, cell in enumerate(row):
            a = _extract_amperage(cell)
            if a and a in FRAME_LADDER:
                found.append((j, a))
        if len(found) >= 3:
            header_row_idx = i
            frame_cols = {j: a for j, a in found}
            break

    if header_row_idx is None:
        return result

    # Scan data rows
    LABEL_MAP = {
        "feeder": ["feeder", "busway feeder", "feeder c/w"],
        "50%e": ["50%e", "50% e", "50%earth"],
        "100%e": ["100%e", "100% e", "100%earth"],
        "flange_end_box": ["flanged end box", "flange end box", "end box (only)"],
        "flange_end": ["flanged end", "flange end"],
        "elbow": ["elbow"],
        "flexible": ["flexible conductor", "flexible link"],
        "hanger_clamp": ["horizontal hanger clamp", "hanger clamp", "mounting clamp"],
        "end_closure": ["end closure"],
        "fixed_hanger": ["vertical fix hanger", "fixed hanger", "fix hanger"],
        "spring_hanger": ["vertical spring hanger", "spring hanger"],
        "plugin_hole": ["plug-in hole", "plugin hole", "plug in hole", "plug-in opening"],
    }

    for row in rows[header_row_idx + 1:]:
        label = str(row[0]).strip().lower() if row else ""

        # Feeder rows carry their earth-% qualifier in the SAME row label
        # (e.g. "Feeder 3P (4W) + 50%E") rather than on a separate sub-row,
        # so both the base rate and the earth-% variant must be detected
        # together here instead of via a two-row lookahead.
        if any(k in label for k in LABEL_MAP["feeder"]):
            # The sheet lists BOTH 3-wire and 4-wire feeder variants for
            # each earth %, in adjacent rows. Mikro quotes 3P4W as standard
            # (BusRun.phases default), so the 3W rows must be skipped —
            # otherwise the later 3W row silently overwrites the correct
            # (higher) 4W price. Same class of collision as the elbow rows.
            compact = label.replace(" ", "")
            if "3w" in compact:
                continue
            has_50 = any(k in label for k in LABEL_MAP["50%e"])
            has_100 = any(k in label for k in LABEL_MAP["100%e"])
            if has_50 or has_100:
                suffix = "50" if has_50 else "100"
                for col, frame_a in frame_cols.items():
                    v = _to_float(row[col] if col < len(row) else None)
                    if v:
                        result[f"feeder_{frame_a}_{suffix}"] = v
            continue

        for key, aliases in LABEL_MAP.items():
            if key in ("feeder", "50%e", "100%e"):
                continue
            # "elbow" needs an exact-label match: the sheet also has
            # "Vertical/Horizontal/Offset/Combination/Special Angle Elbow"
            # rows that all contain the substring "elbow" and would
            # otherwise silently overwrite the plain Elbow rate.
            matched = label == "elbow" if key == "elbow" else any(alias in label for alias in aliases)
            if matched:
                for col, frame_a in frame_cols.items():
                    v = _to_float(row[col] if col < len(row) else None)
                    if v:
                        result[f"{key}_{frame_a}"] = v
                break

    return result


def _parse_piu_sheet(rows: list) -> dict:
    """
    Parse PIU sheet. Rows have MCCB rating; columns have kA tiers.
    Keys: piu_{rating_a}_{ka}
    """
    result: dict = {}
    import re

    # Find header with kA values
    ka_cols: dict[int, int] = {}
    header_idx = None
    for i, row in enumerate(rows):
        found = []
        for j, cell in enumerate(row):
            m = re.search(r"(\d+)\s*k[Aa]", str(cell))
            if m:
                found.append((j, int(m.group(1))))
        if found:
            ka_cols = {j: ka for j, ka in found}
            header_idx = i
            break

    if header_idx is None:
        return result

    # The rating label sits in whichever column precedes the first price
    # column — real sheets have a blank spacer column before it, so
    # column 0 can't be assumed. Range labels like "32A - 100A" must
    # resolve to the upper bound (100), matching PriceList.piu()'s
    # nearest-band-rounded-up fallback lookup.
    min_price_col = min(ka_cols) if ka_cols else 0
    for row in rows[header_idx + 1:]:
        rating = None
        for cell in row[:min_price_col]:
            rating = _extract_upper_amperage(cell)
            if rating:
                break
        if not rating:
            continue
        for col, ka in ka_cols.items():
            v = _to_float(row[col] if col < len(row) else None)
            if v:
                result[f"piu_{rating}_{ka}"] = v

    return result


def _parse_bimetal_sheet(rows: list) -> dict:
    """
    Parse BI METAL PLATE sheet. Rows: frame ratings → price.
    Keys: bimetal_{frame_a}

    Real sheet layout has a blank spacer column before the amperage label,
    and the price (RM/SET) is the last column, with several unrelated
    numeric columns (No/W(mm)/L(mm)) in between — so the label is found by
    scanning the row rather than assuming column 0, and the price is taken
    from the last cell rather than "the first non-empty numeric cell".
    """
    prices: dict = {}
    dims: dict = {}
    for row in rows:
        if not row:
            continue
        idx = None
        a = None
        for i, cell in enumerate(row):
            a = _extract_upper_amperage(cell)
            if a:
                idx = i
                break
        if not a:
            continue
        price = _to_float(row[-1])
        if price:
            prices[f"bimetal_{a}"] = price
        # After the amperage label the numeric cells are No, W(mm), L(mm),
        # then the price. Capture the first three as the plate dimensions.
        nums = [c for c in row[idx + 1:] if isinstance(c, (int, float)) and c]
        if len(nums) >= 4:
            dims[a] = (int(nums[0]), nums[1], nums[2])
    return prices, dims


# Singleton — loaded once at startup or when a new price list is uploaded
price_list = PriceList()
