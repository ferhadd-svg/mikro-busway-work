import csv
from pathlib import Path
from typing import Iterable

import openpyxl
import xlrd

from app.config import settings


SUPPORTED_TEXT_SUFFIXES = {".csv", ".txt", ".xls", ".xlsx"}


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        text = _xlsx_to_text(path)
    elif suffix == ".xls":
        text = _xls_to_text(path)
    elif suffix == ".csv":
        text = _csv_to_text(path)
    elif suffix == ".txt":
        text = path.read_text(encoding="utf-8", errors="replace")
    else:
        raise ValueError(f"Unsupported text extraction file type: {suffix}")

    return text[: settings.max_extracted_text_chars]


def _xlsx_to_text(path: Path) -> str:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    chunks: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        chunks.append(f"Sheet: {sheet_name}")
        chunks.extend(_rows_to_lines(sheet.iter_rows(values_only=True)))
    return "\n".join(chunks)


def _xls_to_text(path: Path) -> str:
    workbook = xlrd.open_workbook(str(path))
    chunks: list[str] = []
    for sheet in workbook.sheets():
        chunks.append(f"Sheet: {sheet.name}")
        for row_index in range(sheet.nrows):
            chunks.append(_clean_row(sheet.row_values(row_index)))
    return "\n".join(chunks)


def _csv_to_text(path: Path) -> str:
    chunks: list[str] = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as file_handle:
        reader = csv.reader(file_handle)
        for row in reader:
            chunks.append(_clean_row(row))
    return "\n".join(chunks)


def _rows_to_lines(rows: Iterable[tuple]) -> list[str]:
    return [_clean_row(row) for row in rows]


def _clean_row(row: Iterable) -> str:
    values = [str(value).strip() for value in row if value not in (None, "")]
    return " | ".join(values)

