"""
Generate the priced quotation Excel by writing item blocks into the
salesperson template (logo, pink header, borders, fonts preserved).

Template structure assumptions (Mikro standard):
  - Header rows: logo, company info, quotation title, client/ref block
  - Item body: starts after a header sentinel row
  - Footer: remarks block + subtotal / SST / grand total + signatures
"""

import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from app.schemas.boq import BOQRun, BOQLineItem, FlagAnswers
from app.models.salesperson import Salesperson
from app.config import settings


REMARKS_AL = [
    "The quantities quoted are rough estimation only. Final busway amount shall based on the actual delivery and approved shop-drawings.",
    "All Bi-Metal materials for Aluminum Busway are by contractor.",
    "Should any modification on factory standard dimensions are chargeable.",
    "All horizontal hanger support are by contractor.",
    "Any delivery to other than project site are subject to additional transportation surcharge.",
    "The prices quoted are EXCLUDING all installation works, all termination of the busway to panels or transformers, testing and commissioning at site.",
    "Warranty against manufacturing defects is 12 calendar months after delivery. This warranty does not cover reimbursement of consequential or incidental damages, labour, transportation, removal of the installation or any other expenses which may be incurred in connection with the repair and replacement.",
]


def _remarks_cu(flags) -> list[str]:
    today = date.today().strftime("%d-%m-%Y")
    return [
        "The quantities quoted are rough estimation only. Final busway amount shall based on the actual delivery and approved shop-drawings.",
        "All Copper Bars are 100% Electro Tin-Plated.",
        f"The prices quoted are based on {today} LME Copper @ USD {flags.lme_usd_per_mt:,.0f}/MT.",
        "Should any modification on factory standard dimensions are chargeable.",
        "All horizontal hanger support are by contractor.",
        "Any delivery to other than project site are subject to additional transportation surcharge.",
        "The prices quoted are EXCLUDING all installation works, all termination of the busway to panels or transformers, testing and commissioning at site.",
        "Warranty against manufacturing defects is 12 calendar months after delivery. This warranty does not cover reimbursement of consequential or incidental damages, labour, transportation, removal of the installation or any other expenses which may be incurred in connection with the repair and replacement.",
    ]


def _terms_block(runs: list, flags) -> list[tuple[str, str]]:
    """Return [(label, value)] rows for Manufacturer/Validity/Delivery/Price/Payment."""
    materials = {r.material for r in runs}
    if "CU" in materials:
        return _terms_cu(flags)
    return _terms_al(flags)


def _terms_al(flags) -> list[tuple[str, str]]:
    validity = (
        f"Prices are based on LME Aluminium at USD {flags.lme_usd_per_mt:,.0f}/MT. "
        f"Any subsequent adjustment shall follow the prevailing LME Aluminium price within a ±2% variation."
    )
    return [
        ("Manufacturer", "Mikro Busway Sdn Bhd, Malaysia."),
        ("Validity",     validity),
        ("Delivery",     "Approximately 8 to 10 working weeks upon receipt of approval drawings."),
        ("Price",        "Ex-Nilai Factory in Ringgit Malaysia (RM)."),
        ("Payment",      "30% Deposit is required upon confirmation of order.\nBalance on Irrevocable Letter of Credit 60 Days."),
    ]


def _terms_cu(flags) -> list[tuple[str, str]]:
    payment = (
        f"30% deposit is required to secure LME CU US{flags.lme_usd_per_mt:,.0f}/MT upon successful transfer to the account \n"
        f"Price within +2% remain unchanged, variations over 2.3% require base of adjustment \n"
        f"The balance is payable via an irrevocable 60days Letter Of Credit "
    )
    return [
        ("Manufacturer", "Mikro Busway Sdn Bhd, Malaysia."),
        ("Validity",     f"Based on LME Copper@USD{flags.lme_usd_per_mt:,.0f}/ MT and thereafter depands on current LME price "),
        ("Delivery",     "Approximately 8 to 10 working weeks upon receipt of approval drawings."),
        ("Price",        "Ex-Nilai Factory in Ringgit Malaysia (RM)."),
        ("Payment",      payment),
        ("Cancellation", "30% from total amount will be imposed on cancellation of purchase order "),
    ]


def _remarks_for_runs(runs: list[BOQRun], flags) -> list[str]:
    materials = {r.material for r in runs}
    if "CU" in materials:
        return _remarks_cu(flags)
    return REMARKS_AL


def build_quotation(
    runs: list[BOQRun],
    flags: FlagAnswers,
    salesperson: Salesperson,
    our_ref: str,
    client_name: str,
    attn: str | None,
    me_consultant: str | None,
    template_path: Path | None,
) -> Path:
    """
    Write a priced quotation Excel. If a salesperson template exists, use it
    as the base; otherwise build from scratch.
    """
    if template_path and template_path.exists():
        wb = openpyxl.load_workbook(str(template_path))
        ws = wb.active
        _fill_template(ws, runs, flags, salesperson, our_ref, client_name, attn, me_consultant)
    else:
        wb = _build_from_scratch(runs, flags, salesperson, our_ref, client_name, attn, me_consultant)

    out_path = settings.projects_dir / f"QUOTATION_{our_ref.replace('/', '-')}_{salesperson.name.replace(' ', '_')}.xlsx"
    wb.save(str(out_path))
    return out_path


# ------------------------------------------------------------------ #
#  Template fill (preferred path)                                     #
# ------------------------------------------------------------------ #

# Column layout of the from-scratch builder; also the fallback when a
# template's item-table header can't be found.
DEFAULT_COLS = {"header_row": 0, "no": 1, "desc": 2, "unit": 3, "qty": 4,
                "rate": 5, "amount": 6}


def _find_header_columns(ws) -> dict | None:
    """Locate the item-table header row (the first row with both a RATE and
    an AMOUNT header) and map out which column holds what. Real templates
    put Quantity/Unit/Unit Rate/Total Amount in different columns than the
    from-scratch layout."""
    for row in ws.iter_rows():
        cells = [(c.column, str(c.value).strip().upper())
                 for c in row if isinstance(c.value, str) and c.value.strip()]
        rate = next((col for col, v in cells if "RATE" in v), None)
        amount = next((col for col, v in cells if "AMOUNT" in v and "RATE" not in v), None)
        if not (rate and amount):
            continue
        return {
            "header_row": row[0].row,
            "no": next((col for col, v in cells if v in ("NO.", "NO", "ITEM", "ITEM NO.")), 1),
            "desc": next((col for col, v in cells if "DESCRIPTION" in v), 2),
            "qty": next((col for col, v in cells if "QUANTITY" in v or "QTY" in v), None),
            "unit": next((col for col, v in cells
                          if v == "UNIT" or (v.startswith("UNIT") and "RATE" not in v)), None),
            "rate": rate,
            "amount": amount,
        }
    return None


def _fill_labelled_fields(ws, salesperson, our_ref, client_name, attn, me_consultant):
    """Real templates label their fields ("To", "Attn", "OUR REF : ") instead
    of using <<tokens>>. Fill each value in after the label's colon."""
    neighbour_labels = {
        "TO": client_name or "",
        "ATTN": attn or "",
        "FROM": salesperson.name,
        "DATE": date.today().strftime("%d-%m-%Y"),
    }
    inline_labels = {
        "OUR REF": our_ref or "",
        "M&E": me_consultant or "",
        "M & E": me_consultant or "",
    }
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip()
            upper = text.upper()

            # e.g. "OUR REF    : " → value appended in the same cell
            for label, value in inline_labels.items():
                if value and upper.startswith(label) and text.endswith(":"):
                    cell.value = cell.value.rstrip() + " " + value
                    break
            else:
                # e.g. A8="To", B8=": " → value written into the next cell
                value = neighbour_labels.get(upper.rstrip(": ").strip())
                if value and upper.rstrip(": ").strip() in neighbour_labels:
                    target = ws.cell(row=cell.row, column=cell.column + 1)
                    existing = str(target.value or "").strip()
                    if existing in ("", ":"):
                        target.value = f": {value}"


def _fill_template(ws, runs, flags, salesperson, our_ref, client_name, attn, me_consultant):
    """
    Fill the salesperson template: header fields, item block, totals.
    Works with both <<token>> templates and label-styled real templates.
    """
    _replace_tokens(ws, {
        "<<OUR_REF>>": our_ref,
        "<<CLIENT>>": client_name or "",
        "<<ATTN>>": attn or "",
        "<<ME>>": me_consultant or "",
        "<<SALESPERSON_NAME>>": salesperson.name,
        "<<SALESPERSON_TITLE>>": salesperson.title,
        "<<SALESPERSON_MOBILE>>": salesperson.mobile,
        "<<SALESPERSON_EMAIL>>": salesperson.email,
        "<<LME_USD>>": f"USD {flags.lme_usd_per_mt:,.0f}/MT",
        "<<USD_MYR>>": f"USD 1 = RM {flags.usd_to_myr:.4f}",
    })
    _fill_labelled_fields(ws, salesperson, our_ref, client_name, attn, me_consultant)

    # Find the row that contains "SUB-TOTAL" or "SUBTOTAL"
    subtotal_row = None
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or "").upper()
            if "SUB-TOTAL" in v or "SUBTOTAL" in v:
                subtotal_row = cell.row
                break
        if subtotal_row:
            break

    if subtotal_row is None:
        # Fallback: append items at the bottom
        _append_items(ws, runs, flags)
        return

    cols = _find_header_columns(ws)

    if cols and cols["header_row"] < subtotal_row:
        # Clear the sample items the template ships with: everything from the
        # first row after the table header that has an item number or a
        # description, down to the row above SUB-TOTAL.
        item_start = None
        for r in range(cols["header_row"] + 1, subtotal_row):
            if (ws.cell(row=r, column=cols["no"]).value not in (None, "")
                    or ws.cell(row=r, column=cols["desc"]).value not in (None, "")):
                item_start = r
                break
        if item_start is not None:
            _delete_rows(ws, item_start, subtotal_row - item_start)
            subtotal_row = item_start
        insert_row = subtotal_row
    else:
        # No recognisable header: first blank row above subtotal
        insert_row = subtotal_row
        for r in range(subtotal_row - 1, 0, -1):
            all_empty = all(
                ws.cell(row=r, column=c).value in (None, "")
                for c in range(1, 7)
            )
            if not all_empty:
                insert_row = r + 1
                break

    sum_rows = _insert_item_block(ws, runs, flags, insert_row, cols)
    _write_totals(ws, runs, insert_row, (cols or {}).get("amount"), sum_rows)


def _shift_images(ws, threshold: int, delta: int):
    """Keep embedded images (company logo, salesperson signature/stamp) anchored
    to the same content as row insertions/deletions happen. openpyxl's
    insert_rows/delete_rows move cell values but never touch image anchors,
    so without this a signature image stays frozen at its original row while
    the "PREPARED BY"/"REVIEWED BY" text it belongs next to moves — leaving
    it floating over the wrong content."""
    for img in getattr(ws, "_images", []):
        anchor = img.anchor
        frm = getattr(anchor, "_from", None)
        if frm is not None and frm.row >= threshold:
            frm.row += delta
        to = getattr(anchor, "to", None)
        if to is not None and to.row >= threshold:
            to.row += delta


def _insert_row(ws, row: int):
    """Insert one row, keeping merged cells and embedded images aligned with
    their values.

    openpyxl's insert_rows shifts cell values down but leaves merged ranges
    (and image anchors) at their old coordinates, so every merge or image at
    or below the insertion point (the totals block, remarks, signature area
    at the end of the template) drifts one row out of place per inserted
    item. Re-anchor them manually.
    """
    to_shift, to_expand = [], []
    for rng in ws.merged_cells.ranges:
        if rng.min_row >= row:
            to_shift.append(str(rng))
        elif rng.max_row >= row:
            to_expand.append(str(rng))
    for ref in to_shift + to_expand:
        ws.unmerge_cells(ref)
    ws.insert_rows(row)
    for ref in to_shift:
        r = CellRange(ref)
        r.shift(row_shift=1)
        ws.merge_cells(str(r))
    for ref in to_expand:
        r = CellRange(ref)
        r.expand(down=1)
        ws.merge_cells(str(r))
    _shift_images(ws, row, 1)


def _delete_rows(ws, idx: int, amount: int):
    """Delete rows with the same merged-range and image-anchor bookkeeping as
    _insert_row: ranges/images below the deleted span shift up, ranges
    overlapping it are unmerged (their rows are gone)."""
    last = idx + amount - 1
    to_shift, to_drop = [], []
    for rng in ws.merged_cells.ranges:
        if rng.min_row > last:
            to_shift.append(str(rng))
        elif rng.max_row >= idx:
            to_drop.append(str(rng))
    for ref in to_shift + to_drop:
        ws.unmerge_cells(ref)
    ws.delete_rows(idx, amount)
    for ref in to_shift:
        r = CellRange(ref)
        r.shift(row_shift=-amount)
        ws.merge_cells(str(r))
    _shift_images(ws, last + 1, -amount)


def _replace_tokens(ws, token_map: dict):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for token, value in token_map.items():
                    if token in cell.value:
                        cell.value = cell.value.replace(token, value)


def _run_title(run: BOQRun) -> str:
    """Mikro house-format spec title, e.g.
    'MIKRO BUSWAY # 5000A TPNE, 3P4W+100%E, 600VAC, 50Hz (COPPER) - IP54'."""
    material = "ALUMINIUM" if run.material == "AL" else "COPPER"
    if run.frame_rating_a and run.earth_pct:
        return (f"MIKRO BUSWAY # {run.frame_rating_a}A TPNE, "
                f"{run.phases}+{run.earth_pct}%E, 600VAC, 50Hz ({material}) - IP54")
    return f"MIKRO BUSWAY — {run.run_id} ({run.run_type}, {material})"


def _short_desc(desc: str, run: BOQRun | None) -> str:
    """Component descriptions are now emitted in final house wording by
    boq_builder (no frame-rating suffix to strip), so this is a pass-through
    kept only so callers don't need to change."""
    return desc


def _border_row(ws, row: int, cols: dict, border):
    for col in range(1, cols["amount"] + 1):
        ws.cell(row=row, column=col).border = border


def _insert_item_block(ws, runs: list[BOQRun], flags: FlagAnswers, start_row: int,
                       cols: dict | None = None) -> list[int]:
    """Write the runs in the Mikro house format: item number = run number,
    spec-title row, routing row, component lines with =qty*rate formulas,
    then a per-run =SUM() amount row. Returns the per-run sum rows so the
    totals block can reference them."""
    cols = cols or DEFAULT_COLS
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    amount_letter = get_column_letter(cols["amount"])
    bold = Font(bold=True)

    row = start_row
    sum_rows: list[int] = []

    for n, run in enumerate(runs, 1):
        # Item number + spec title
        _insert_row(ws, row)
        _border_row(ws, row, cols, border)
        ws.cell(row=row, column=cols["no"], value=n).font = bold
        ws.cell(row=row, column=cols["desc"], value=_run_title(run)).font = bold
        row += 1

        # Routing line
        _insert_row(ws, row)
        _border_row(ws, row, cols, border)
        ws.cell(row=row, column=cols["desc"], value=run.routing or "").font = bold
        row += 1

        first_amount_row = row
        for item in run.items:
            _insert_row(ws, row)
            _write_quotation_row(ws, row, item, border, cols, run)
            row += 1

        if run.piu_items:
            _insert_row(ws, row)
            _border_row(ws, row, cols, border)
            c = ws.cell(row=row, column=cols["desc"], value="PLUG-IN UNITS (PIU) :")
            c.font = Font(italic=True, bold=True)
            row += 1
            for item in run.piu_items:
                _insert_row(ws, row)
                _write_quotation_row(ws, row, item, border, cols, run)
                row += 1

        # Per-run amount subtotal (house format: =SUM over the run's lines)
        _insert_row(ws, row)
        _border_row(ws, row, cols, border)
        c = ws.cell(row=row, column=cols["amount"],
                    value=f"=SUM({amount_letter}{first_amount_row}:{amount_letter}{row - 1})")
        c.font = bold
        sum_rows.append(row)
        row += 1

    return sum_rows


def _write_quotation_row(ws, row: int, item: BOQLineItem, border,
                         cols: dict | None = None, run: BOQRun | None = None):
    cols = cols or DEFAULT_COLS

    # OPTIONAL subheader: bold label only, no quantity/price.
    if getattr(item, "is_subheader", False):
        _border_row(ws, row, cols, border)
        ws.cell(row=row, column=cols["desc"], value=item.description).font = Font(bold=True)
        return

    # Quoted-out line (Connection Bars): literal "EXCLUDED", no formula.
    if getattr(item, "is_excluded", False):
        _border_row(ws, row, cols, border)
        ws.cell(row=row, column=cols["desc"], value=item.description)
        if cols.get("unit"):
            ws.cell(row=row, column=cols["unit"], value=item.unit)
        for key in ("rate", "amount"):
            c = ws.cell(row=row, column=cols[key], value="EXCLUDED")
            c.alignment = Alignment(horizontal="right")
        return

    mapping = [("desc", _short_desc(item.description, run)), ("unit", item.unit),
               ("qty", item.qty), ("rate", item.unit_rate_myr)]
    values = {cols[key]: val for key, val in mapping if cols.get(key)}
    if cols.get("qty"):
        qty_l, rate_l = get_column_letter(cols["qty"]), get_column_letter(cols["rate"])
        values[cols["amount"]] = f"={qty_l}{row}*{rate_l}{row}"
    else:
        values[cols["amount"]] = item.amount_myr
    right_cols = {cols.get("qty"), cols["rate"], cols["amount"]}
    for col in range(1, cols["amount"] + 1):
        c = ws.cell(row=row, column=col, value=values.get(col))
        c.border = border
        if col in right_cols:
            c.alignment = Alignment(horizontal="right")


def _subtotal_sst_grand(runs: list[BOQRun]) -> tuple[int, int, int]:
    subtotal = round(sum(
        item.amount_myr for r in runs for item in (r.items + r.piu_items)
    ))
    sst = round(subtotal * 0.10)
    return subtotal, sst, subtotal + sst


def _write_totals(ws, runs: list[BOQRun], min_row: int, amount_col: int | None = None,
                  sum_rows: list[int] | None = None):
    subtotal, sst, grand_total = _subtotal_sst_grand(runs)
    # Write into the AMOUNT column when the template has one; otherwise fall
    # back to the historical assumption of two columns right of the label.
    if amount_col is None:
        amount_col = (_find_header_columns(ws) or {}).get("amount")

    # Locate the three label cells first
    sub_cell = sst_cell = grand_cell = None
    for row in ws.iter_rows(min_row=min_row):
        for cell in row:
            v = str(cell.value or "").upper()
            if ("SUB-TOTAL" in v or "SUBTOTAL" in v) and sub_cell is None:
                sub_cell = cell
            elif "GRAND TOTAL" in v and grand_cell is None:
                grand_cell = cell
            elif ("SST" in v or "TAX" in v) and sst_cell is None:
                sst_cell = cell

    def _target(label_cell):
        return ws.cell(row=label_cell.row, column=amount_col or (label_cell.column + 2))

    # House format: live formulas chained off the per-run sum rows.
    # Without a known amount column, fall back to literal values.
    use_formulas = amount_col and sum_rows and sub_cell
    letter = get_column_letter(amount_col) if amount_col else None

    if sub_cell:
        _target(sub_cell).value = (
            f"=SUM({','.join(f'{letter}{r}' for r in sum_rows)})" if use_formulas else subtotal
        )
    if sst_cell:
        _target(sst_cell).value = (
            f"={letter}{sub_cell.row}*10%" if use_formulas and sst_cell else sst
        )
    if grand_cell:
        _target(grand_cell).value = (
            f"=SUM({letter}{sub_cell.row},{letter}{sst_cell.row})"
            if use_formulas and sst_cell else grand_total
        )
        # Refresh "Item No. 1 to N" in the grand-total label
        if isinstance(grand_cell.value, str):
            n = len(runs)
            replacement = "Item No. 1" if n == 1 else f"Item No. 1 to {n}"
            grand_cell.value = re.sub(r"Item No\.?\s*1(\s*to\s*\d+)?", replacement,
                                      grand_cell.value)


def _append_items(ws, runs: list[BOQRun], flags: FlagAnswers):
    """Last-resort: write items at first empty row, then a totals block."""
    max_row = ws.max_row + 2
    _insert_item_block(ws, runs, flags, max_row)

    subtotal, sst, grand_total = _subtotal_sst_grand(runs)
    bold = Font(bold=True)
    row = ws.max_row + 1
    for label, value in [("SUB-TOTAL (RM)", subtotal),
                         ("10% SST (RM)", sst),
                         ("GRAND TOTAL (RM)", grand_total)]:
        ws.cell(row=row, column=5, value=label).font = bold
        ws.cell(row=row, column=6, value=value).font = bold
        row += 1


# ------------------------------------------------------------------ #
#  From-scratch builder (no template)                                 #
# ------------------------------------------------------------------ #

def _build_from_scratch(runs, flags, salesperson, our_ref, client_name, attn, me_consultant):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QUOTATION"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    bold = Font(bold=True)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "MIKRO ENGINEERING SDN BHD"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "BUSWAY SYSTEM QUOTATION"
    ws["A2"].font = header_font
    ws["A2"].fill = header_fill
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A4"] = "Our Ref:"
    ws["B4"] = our_ref
    ws["A5"] = "Client:"
    ws["B5"] = client_name
    if attn:
        ws["A6"] = "Attn:"
        ws["B6"] = attn
    if me_consultant:
        ws["A7"] = "M&E:"
        ws["B7"] = me_consultant

    ws["D4"] = "Salesperson:"
    ws["E4"] = salesperson.name
    ws["D5"] = "Title:"
    ws["E5"] = salesperson.title
    ws["D6"] = "Mobile:"
    ws["E6"] = salesperson.mobile

    # Terms block (Manufacturer, Validity, Delivery, Price, Payment)
    row = 10
    label_font = Font(bold=True)
    for label, value in _terms_block(runs, flags):
        ws.cell(row=row, column=1, value=label).font = label_font
        c = ws.cell(row=row, column=2, value=value)
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30 if "\n" in value or len(value) > 80 else 15
        row += 1

    row += 1  # blank spacer

    for col, header in enumerate(["No.", "DESCRIPTION", "UNIT", "QTY", "UNIT RATE (RM)", "AMOUNT (RM)"], 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row += 1

    _insert_item_block(ws, runs, flags, row)

    subtotal, sst, grand = _subtotal_sst_grand(runs)
    end_row = ws.max_row + 2

    ws.cell(row=end_row, column=5, value="SUB-TOTAL (RM)").font = bold
    ws.cell(row=end_row, column=6, value=subtotal).font = bold
    end_row += 1
    ws.cell(row=end_row, column=5, value="10% SST (RM)").font = bold
    ws.cell(row=end_row, column=6, value=sst).font = bold
    end_row += 1
    ws.cell(row=end_row, column=5, value="GRAND TOTAL (RM)").font = Font(bold=True, size=12)
    ws.cell(row=end_row, column=6, value=grand).font = Font(bold=True, size=12)

    end_row += 2
    remarks = _remarks_for_runs(runs, flags)
    ws.cell(row=end_row, column=1, value="Remarks :").font = Font(bold=True, underline="single")
    for i, remark in enumerate(remarks, 1):
        c = ws.cell(row=end_row + i, column=1, value=f"{i}.  {remark}")
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[end_row + i].height = 30 if len(remark) > 100 else 15

    sig_row = end_row + len(remarks) + 3
    ws.cell(row=sig_row, column=1, value="Prepared by:").font = bold
    ws.cell(row=sig_row + 2, column=1, value=salesperson.name)
    ws.cell(row=sig_row + 3, column=1, value=salesperson.title)
    ws.cell(row=sig_row + 4, column=1, value=f"Tel: {salesperson.mobile}")
    ws.cell(row=sig_row + 5, column=1, value=f"Email: {salesperson.email}")

    return wb
