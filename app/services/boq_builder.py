"""
Build the Bill of Quantities from extracted drawing data + flag answers.

Accessory templates per run type (from Mikro Busway rules):

TX-MSB:      Feeder, Flange End, H.Elbow, V.Elbow, Flexible Link, [opt] Mounting Clamp + Bi-metal
             Connection Bars EXCLUDED.

MSB-Riser:   Feeder, Flange End, End Closure, H/V Elbow, Fixed Hanger, Spring Hanger,
             Plug-in Opening, [opt] Spare + Bi-metal, then PIU section.

RISER:       Feeder, Cable Entry Box, End Closure, Fixed Hanger, Spring Hanger,
             Plug-in Opening, [opt] Spare + Bi-metal, then PIU section.
"""

import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.boq import (
    BusRun, DrawingExtraction, FlagAnswers, BOQLineItem, BOQRun, BOQResponse
)
from app.services.price_list import price_list, resolve_frame_rating
from app.config import settings


# ------------------------------------------------------------------ #
#  Hanger quantity helper                                             #
# ------------------------------------------------------------------ #

def _calc_hangers(length_m: float, spacing_m: float = 1.5) -> tuple[int, int]:
    """
    Return (fixed, spring) hanger counts.
    Typical pattern: fix, fix, spring repeating every 3 hangers.
    Minimum 2 hangers per run.
    """
    total = max(2, math.ceil(length_m / spacing_m) + 1)
    spring = max(1, total // 3)
    fixed = total - spring
    return fixed, spring


# ------------------------------------------------------------------ #
#  Line-item builders per run type                                    #
# ------------------------------------------------------------------ #

def _line(description: str, unit: str, qty: float, rate: float) -> BOQLineItem:
    return BOQLineItem(
        description=description,
        unit=unit,
        qty=qty,
        unit_rate_myr=round(rate),
        amount_myr=round(qty * round(rate)),
    )


def _subheader(label: str = "OPTIONAL") -> BOQLineItem:
    """A label-only row (e.g. 'OPTIONAL') that carries no price."""
    return BOQLineItem(description=label, unit="", qty=0, unit_rate_myr=0,
                       amount_myr=0, is_subheader=True)


def _excluded(description: str, unit: str = "LOTS") -> BOQLineItem:
    """A line that is quoted out — shows 'EXCLUDED' instead of a price
    (house format for 'CONNECTION BARS (TX & MSB)')."""
    return BOQLineItem(description=description, unit=unit, qty=1, unit_rate_myr=0,
                       amount_myr=0, is_excluded=True)


# House-format line descriptions / units (must match the salesperson template
# and MK.451 verbatim — the run title already carries rating/earth/material,
# so component lines do NOT repeat the frame rating).
_FEEDER_DESC = "FEEDER C/W INTEGRAL EARTH  (Horizontal)"
_U_MTR, _U_NOS, _U_SETS, _U_LOTS = "MTR", "NOS", "SETS", "LOTS"


def _mounting_clamp_qty(length_m: float) -> int:
    """House estimate: ~1 clamp per 2m (MK.451: 10m run → 5 clamps)."""
    return max(2, math.ceil((length_m or 0.0) / 2.0))


def _build_tx_msb(run: BusRun) -> list[BOQLineItem]:
    """TX-MSB accessory template (skill spec):
    Feeder · Flange End · Horizontal Elbow · Vertical Elbow · Flexible Link
    · OPTIONAL[ Mounting Clamp · Bi-metal(Al) ] · Connection Bars=EXCLUDED."""
    fa = run.frame_rating_a
    m = run.material
    length = run.length_m or 0.0
    items: list[BOQLineItem] = [
        _line(_FEEDER_DESC, _U_MTR, length, price_list.feeder(fa, run.earth_pct, m)),
        _line("FLANGE END", _U_NOS, 2, price_list.flange_end(fa, m)),
        _line("HORIZONTAL ELBOW", _U_NOS, 2, price_list.elbow(fa, m)),
        _line("VERTICAL ELBOW", _U_NOS, 2, price_list.vertical_elbow(fa, m)),
        _line("FLEXIBLE LINK (BRAIDED TYPE)", _U_SETS, 1, price_list.flexible_conductor(fa, m)),
        _subheader("OPTIONAL"),
        _line("MOUNTING CLAMP (W/O ROD & C-CHANNEL)", _U_SETS,
              _mounting_clamp_qty(length), price_list.mounting_clamp(fa, m)),
    ]
    if m == "AL":
        items.append(_line("BI-METAL PLATE", _U_SETS, 2, price_list.bimetal(fa)))
    items.append(_excluded("CONNECTION BARS (TX & MSB)", _U_LOTS))
    return items


def _build_msb_riser(run: BusRun, piu_ka: int) -> tuple[list[BOQLineItem], list[BOQLineItem]]:
    """MSB-Riser accessory template (skill spec):
    Feeder · Flange End · End Closure · Horizontal Elbow · Vertical Elbow
    · Fixed Hanger · Spring Hanger · Plug-in Opening
    · OPTIONAL[ Plug-in Opening (Spare) · Bi-metal(Al) · Mounting Clamp ] · PIU."""
    fa = run.frame_rating_a
    m = run.material
    length = run.length_m or 0.0

    fixed, spring = (run.num_fixed_hangers, run.num_spring_hangers)
    if fixed is None or spring is None:
        fixed, spring = _calc_hangers(length, run.hanger_spacing_m)

    items: list[BOQLineItem] = [
        _line(_FEEDER_DESC, _U_MTR, length, price_list.feeder(fa, run.earth_pct, m)),
        _line("FLANGE END", _U_NOS, 2, price_list.flange_end(fa, m)),
        _line("END CLOSURE", _U_NOS, 1, price_list.end_closure(fa, m)),
        _line("HORIZONTAL ELBOW", _U_NOS, 2, price_list.elbow(fa, m)),
        _line("VERTICAL ELBOW", _U_NOS, 2, price_list.vertical_elbow(fa, m)),
        _line("FIXED HANGER", _U_SETS, fixed, price_list.fixed_hanger(fa, m)),
        _line("SPRING HANGER", _U_SETS, spring, price_list.spring_hanger(fa, m)),
    ]
    if len(run.piu_ratings) > 0:
        items.append(_line("PLUG-IN OPENING", _U_NOS, len(run.piu_ratings),
                           price_list.plugin_opening(fa, m)))

    optional: list[BOQLineItem] = []
    if run.spare_openings > 0:
        optional.append(_line("PLUG-IN OPENING (SPARE)", _U_NOS, run.spare_openings,
                              price_list.plugin_opening(fa, m)))
    optional.append(_line("MOUNTING CLAMP (W/O ROD & C-CHANNEL)", _U_SETS,
                          _mounting_clamp_qty(length), price_list.mounting_clamp(fa, m)))
    if m == "AL":
        optional.append(_line("BI-METAL PLATE", _U_SETS, 2, price_list.bimetal(fa)))
    if optional:
        items.append(_subheader("OPTIONAL"))
        items.extend(optional)

    return items, _build_piu(run, piu_ka)


def _build_riser(run: BusRun, piu_ka: int) -> tuple[list[BOQLineItem], list[BOQLineItem]]:
    """RISER (cable-entry) accessory template (skill spec):
    Feeder · Cable Entry Box · End Closure · Fixed Hanger · Spring Hanger
    · Plug-in Opening · OPTIONAL[ Plug-in Opening (Spare) · Bi-metal(Al) ] · PIU."""
    fa = run.frame_rating_a
    m = run.material
    length = run.length_m or 0.0

    fixed, spring = (run.num_fixed_hangers, run.num_spring_hangers)
    if fixed is None or spring is None:
        fixed, spring = _calc_hangers(length, run.hanger_spacing_m)

    items: list[BOQLineItem] = [
        _line(_FEEDER_DESC, _U_MTR, length, price_list.feeder(fa, run.earth_pct, m)),
        _line("CABLE ENTRY BOX", _U_NOS, 1, price_list.cable_entry_box(fa, m)),
        _line("END CLOSURE", _U_NOS, 1, price_list.end_closure(fa, m)),
        _line("FIXED HANGER", _U_SETS, fixed, price_list.fixed_hanger(fa, m)),
        _line("SPRING HANGER", _U_SETS, spring, price_list.spring_hanger(fa, m)),
    ]
    if len(run.piu_ratings) > 0:
        items.append(_line("PLUG-IN OPENING", _U_NOS, len(run.piu_ratings),
                           price_list.plugin_opening(fa, m)))

    optional: list[BOQLineItem] = []
    if run.spare_openings > 0:
        optional.append(_line("PLUG-IN OPENING (SPARE)", _U_NOS, run.spare_openings,
                              price_list.plugin_opening(fa, m)))
    if m == "AL":
        optional.append(_line("BI-METAL PLATE", _U_SETS, 2, price_list.bimetal(fa)))
    if optional:
        items.append(_subheader("OPTIONAL"))
        items.extend(optional)

    return items, _build_piu(run, piu_ka)


def _build_piu(run: BusRun, ka: int) -> list[BOQLineItem]:
    piu_items: list[BOQLineItem] = []
    # Group duplicate ratings
    from collections import Counter
    counts = Counter(run.piu_ratings)
    for rating_a, qty in sorted(counts.items()):
        piu_items.append(_line(
            f"{rating_a}A TPN MCCB (HYUNDAI)",
            _U_NOS, qty, price_list.piu(rating_a, ka)
        ))
    return piu_items


# ------------------------------------------------------------------ #
#  Main BOQ builder                                                   #
# ------------------------------------------------------------------ #

def build_boq(
    extraction: DrawingExtraction,
    flags: FlagAnswers,
    our_ref: str,
    client_name: str,
) -> BOQResponse:
    boq_runs: list[BOQRun] = []

    for run in extraction.runs:
        # Apply any per-run overrides from flag answers
        overrides = flags.run_overrides.get(run.run_id, {})
        if overrides:
            if "rating_a" in overrides and "frame_rating_a" not in overrides:
                overrides = {**overrides, "frame_rating_a": resolve_frame_rating(overrides["rating_a"])}
            run = run.model_copy(update=overrides)

        if run.run_type == "TX-MSB":
            items = _build_tx_msb(run)
            piu_items: list[BOQLineItem] = []
        elif run.run_type == "MSB-Riser":
            items, piu_items = _build_msb_riser(run, flags.piu_ka)
        else:  # RISER
            items, piu_items = _build_riser(run, flags.piu_ka)

        boq_runs.append(BOQRun(
            run_id=run.run_id,
            routing=run.routing,
            run_type=run.run_type,
            material=run.material,
            items=items,
            piu_items=piu_items,
            frame_rating_a=run.frame_rating_a,
            earth_pct=run.earth_pct,
            phases=run.phases,
        ))

    subtotal = sum(
        item.amount_myr
        for r in boq_runs
        for item in (r.items + r.piu_items)
    )

    boq_file = _write_boq_excel(boq_runs, our_ref, client_name, subtotal)

    return BOQResponse(
        project_our_ref=our_ref,
        runs=boq_runs,
        subtotal_myr=subtotal,
        boq_file=str(boq_file),
    )


# ------------------------------------------------------------------ #
#  Excel writer                                                       #
# ------------------------------------------------------------------ #

def _write_boq_excel(runs: list[BOQRun], our_ref: str, client_name: str, subtotal: float) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    section_font = Font(bold=True)
    bold = Font(bold=True)

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "BILL OF QUANTITIES — BUSWAY SYSTEM"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Project Ref: {our_ref}    Client: {client_name}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    row = 4
    for col, header in enumerate(["DESCRIPTION", "UNIT", "QTY", "UNIT RATE (RM)", "AMOUNT (RM)"], 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    row += 1
    grand_total_formula_rows = []

    for run in runs:
        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1,
                    value=f"{run.run_id} — {run.routing} ({run.run_type}, {run.material})")
        c.font = section_font
        c.fill = section_fill
        c.border = border
        row += 1

        for item in run.items:
            _write_item_row(ws, row, item, border)
            grand_total_formula_rows.append(row)
            row += 1

        if run.piu_items:
            ws.merge_cells(f"A{row}:E{row}")
            c = ws.cell(row=row, column=1, value="PIU — PLUG-IN UNITS")
            c.font = Font(italic=True, bold=True)
            c.border = border
            row += 1
            for item in run.piu_items:
                _write_item_row(ws, row, item, border)
                grand_total_formula_rows.append(row)
                row += 1

        row += 1  # blank row between runs

    # Subtotal
    ws.cell(row=row, column=4, value="SUB-TOTAL (RM)").font = bold
    ws.cell(row=row, column=5, value=round(subtotal)).font = bold
    row += 1
    sst = round(subtotal * 0.10)
    ws.cell(row=row, column=4, value="10% SST (RM)").font = bold
    ws.cell(row=row, column=5, value=sst).font = bold
    row += 1
    ws.cell(row=row, column=4, value="GRAND TOTAL (RM)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=5, value=round(subtotal + sst)).font = Font(bold=True, size=12)

    out_path = settings.projects_dir / f"BOQ_{our_ref.replace('/', '-')}.xlsx"
    wb.save(str(out_path))
    return out_path


def _write_item_row(ws, row: int, item: BOQLineItem, border) -> None:
    if item.is_subheader:
        c = ws.cell(row=row, column=1, value=item.description)
        c.font = Font(italic=True, bold=True)
        c.border = border
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = border
        return
    if item.is_excluded:
        values = [item.description, item.unit, "", "EXCLUDED", "EXCLUDED"]
    else:
        values = [item.description, item.unit, item.qty, item.unit_rate_myr, item.amount_myr]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border
        if col >= 3:
            c.alignment = Alignment(horizontal="right")
